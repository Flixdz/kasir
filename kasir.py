import streamlit as st
import openpyxl
from openpyxl import Workbook
import pandas as pd  # Pastikan Anda mengimpor pandas untuk menggunakan DataFrame

# Fungsi untuk membaca data barang dari file Excel
def baca_dari_excel(nama_file):
    try:
        wb = openpyxl.load_workbook(nama_file)
        ws = wb.active
        
        # menyimpan barang
        barang = {}
        
        # Membaca data dari setiap baris
        for row in range(2, ws.max_row + 1):
            id_barang = ws.cell(row=row, column=1).value
            nama_barang = ws.cell(row=row, column=2).value
            harga_barang = ws.cell(row=row, column=3).value
            stok_barang = ws.cell(row=row, column=4).value
            
            barang[id_barang] = {
                'nama': nama_barang,
                'harga': harga_barang,
                'stok': stok_barang
            }
        
        return barang
        
    except FileNotFoundError:
        st.error(f"File '{nama_file}' tidak ditemukan. Membuat dictionary barang kosong.")
        return {}

# Fungsi untuk menyimpan data barang ke dalam file Excel
def simpan_ke_excel(barang, nama_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Barang"
    
    # Menulis header kolom
    ws['A1'] = "ID"
    ws['B1'] = "Nama Barang"
    ws['C1'] = "Harga"
    ws['D1'] = "Stok"
    
    # Menulis data barang
    row = 2
    for id_barang, detail_barang in barang.items():
        ws.cell(row=row, column=1).value = id_barang
        ws.cell(row=row, column=2).value = detail_barang['nama']
        ws.cell(row=row, column=3).value = detail_barang['harga']
        ws.cell(row=row, column=4).value = detail_barang['stok']
        row += 1
    
    # Menyimpan ke file Excel
    wb.save(nama_file)
    st.success(f"Data barang berhasil disimpan ke {nama_file}")

# Fungsi untuk menampilkan semua barang di Streamlit
def tampilkan_semua_barang(barang):
    st.subheader("Daftar Barang")
    
    # Membuat data frame dari dictionary barang
    data_barang = {
        "ID": [],
        "Nama Barang": [],
        "Harga": [],
        "Stok": []
    }
    
    for id_barang, detail_barang in barang.items():
        data_barang["ID"].append(id_barang)
        data_barang["Nama Barang"].append(detail_barang['nama'])
        data_barang["Harga"].append(detail_barang['harga'])
        data_barang["Stok"].append(detail_barang['stok'])
    
    # Menampilkan data barang dalam format tabel
    df_barang = pd.DataFrame(data_barang)
    st.dataframe(df_barang, use_container_width=True)

# Fungsi untuk menambah barang baru
def tambah_barang(barang):
    st.subheader("Tambah Barang Baru")
    nama = st.text_input("Masukkan nama barang:")
    harga = st.number_input("Masukkan harga barang:", min_value=0)
    stok = st.number_input("Masukkan stok barang:", min_value=0, step=1)
    
    if st.button("Tambah"):
        id_barang_baru = max(barang.keys(), default=0) + 1
        barang[id_barang_baru] = {'nama': nama, 'harga': harga, 'stok': stok}
        simpan_ke_excel(barang, "data_barang.xlsx")  # Simpan data barang yang sudah diupdate ke file Excel
        st.success(f"Barang '{nama}' dengan ID {id_barang_baru} berhasil ditambahkan.")
        st.session_state.barang = barang  # Update session state

# Fungsi untuk mencari barang berdasarkan ID
def cari_barang(barang):
    st.subheader("Cari Barang")
    id_barang = st.number_input("Masukkan ID barang yang ingin dicari:", min_value=1, step=1)
    
    if st.button("Cari"):
        if id_barang in barang:
            detail = barang[id_barang]
            st.write(f"ID: {id_barang}")
            st.write(f"Nama Barang: {detail['nama']}")
            st.write(f"Harga: {detail['harga']}")
            st.write(f"Stok: {detail['stok']}")
        else:
            st.error(f"Barang dengan ID {id_barang} tidak ditemukan.")

# Fungsi untuk memodifikasi atau menghapus barang
def modifikasi_barang(barang):
    st.subheader("Modifikasi Barang")
    
    id_barang = st.number_input("Masukkan ID barang yang ingin dimodifikasi atau dihapus:", min_value=1, step=1)
    
    if id_barang in barang:
        nama_baru = st.text_input("Masukkan nama baru:", value=barang[id_barang]['nama'])
        harga_baru = st.number_input("Masukkan harga baru:", min_value=0, value=barang[id_barang]['harga'])
        stok_baru = st.number_input("Masukkan stok baru:", min_value=0, step=1, value=barang[id_barang]['stok'])
        
        if st.button("Simpan Perubahan"):
            barang[id_barang]['nama'] = nama_baru
            barang[id_barang]['harga'] = harga_baru
            barang[id_barang]['stok'] = stok_baru
            simpan_ke_excel(barang, "data_barang.xlsx")  # Simpan data barang yang sudah diupdate ke file Excel
            st.success(f"Barang dengan ID {id_barang} berhasil dimodifikasi.")
            st.session_state.barang = barang  # Update session state
        
        if st.button("Hapus Barang"):
            del barang[id_barang]
            simpan_ke_excel(barang, "data_barang.xlsx")  # Simpan data barang yang sudah diupdate ke file Excel
            st.success(f"Barang dengan ID {id_barang} berhasil dihapus.")
            st.session_state.barang = barang  # Update session state
    else:
        st.error(f"Barang dengan ID {id_barang} tidak ditemukan.")

# Inisialisasi counter transaksi
counter_transaksi = 1

# Fungsi untuk melakukan pembelian barang
def beli_barang(barang):
    global counter_transaksi
    st.subheader("Beli Barang")
    total_belanja = 0
    transaksi = []  # untuk menyimpan detail transaksi
    
    # List untuk menyimpan ID barang yang ingin dibeli
    list_id_barang = []
    # Dict untuk menyimpan jumlah barang yang ingin dibeli dengan key ID barang
    jumlah_barang_dibeli = {}
    
    barang_list = [f"{barang[id_barang]['nama']} (ID: {id_barang})" for id_barang in barang]
    
    while True:
        selected_barang = st.selectbox("Pilih barang yang ingin dibeli:", [""] + barang_list, key=f"select_barang_{counter_transaksi}")
        
        if selected_barang == "":
            break  # keluar dari loop jika pengguna memilih untuk selesai
        
        id_barang = int(selected_barang.split("(ID: ")[1][:-1])
        
        if id_barang in barang:
            if id_barang not in list_id_barang:
                list_id_barang.append(id_barang)
                jumlah_barang_dibeli[id_barang] = 1  # Set jumlah awal untuk barang yang baru diinput
            
            jumlah = st.number_input(f"Masukkan jumlah {barang[id_barang]['nama']} yang ingin dibeli:", min_value=0, step=1, value=jumlah_barang_dibeli[id_barang], key=f"jumlah_input_{counter_transaksi}_{id_barang}")
            jumlah_barang_dibeli[id_barang] = jumlah  # Update jumlah barang yang ingin dibeli
                
        else:
            st.error(f"Barang dengan ID {id_barang} tidak ditemukan.")
        
        # Update counter transaksi
        counter_transaksi += 1
    
    # Tampilkan daftar barang yang akan dibeli
    if list_id_barang:
        st.subheader("Daftar Barang yang Akan Dibeli")
        for id_barang in list_id_barang:
            st.write(f"{barang[id_barang]['nama']} (ID: {id_barang}) - Jumlah: {jumlah_barang_dibeli[id_barang]} - Harga Satuan: Rp {barang[id_barang]['harga']} - Total Harga: Rp {barang[id_barang]['harga'] * jumlah_barang_dibeli[id_barang]}")
            total_belanja += barang[id_barang]['harga'] * jumlah_barang_dibeli[id_barang]
        
        st.write(f"Total belanjaan adalah Rp {total_belanja}")
        
        if st.button("Proses Pembelian"):
            for id_barang in list_id_barang:
                jumlah = jumlah_barang_dibeli[id_barang]
                total_harga = barang[id_barang]['harga'] * jumlah
                
                # Simpan detail transaksi
                transaksi.append({
                    'ID Barang': id_barang,
                    'Nama Barang': barang[id_barang]['nama'],
                    'Harga Satuan': barang[id_barang]['harga'],
                    'Jumlah': jumlah,
                    'Total Harga': total_harga
                })
                
                barang[id_barang]['stok'] -= jumlah  # Update stok barang
            
            cetak_bukti_pembayaran(transaksi)
            simpan_ke_excel(barang, "data_barang.xlsx")  # Simpan data barang yang sudah diupdate ke file Excel
            st.success("Pembelian berhasil diproses!")
            # Kosongkan list transaksi setelah mencetak bukti pembayaran
            transaksi = []
    else:
        st.warning("Belum ada barang yang dipilih.")

# Fungsi untuk mencetak bukti pembayaran
def cetak_bukti_pembayaran(transaksi):
    st.subheader("Bukti Pembayaran")
    df_transaksi = pd.DataFrame(transaksi)
    st.table(df_transaksi)
    st.write(f"Total Pembayaran: Rp {df_transaksi['Total Harga'].sum()}")
    
    # Menyimpan ke file Excel sementara
    nama_file_bukti = "bukti_pembayaran.xlsx"
    df_transaksi.to_excel(nama_file_bukti, index=False)
    
    # Menyediakan link download untuk bukti pembayaran
    with open(nama_file_bukti, "rb") as f:
        st.download_button(
            label="Unduh Bukti Pembayaran",
            data=f,
            file_name=nama_file_bukti,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# Fungsi untuk menampilkan menu utama
def tampilkan_menu(barang):
    st.sidebar.title("Menu")
    pilihan_menu = st.sidebar.selectbox("Pilih menu:", ["Tampilkan Semua Barang", "Tambah Barang Baru", "Cari Barang", "Beli Barang", "Modifikasi Barang", "Keluar"])
    
    if pilihan_menu == "Tampilkan Semua Barang":
        tampilkan_semua_barang(barang)
    elif pilihan_menu == "Tambah Barang Baru":
        tambah_barang(barang)
    elif pilihan_menu == "Cari Barang":
        cari_barang(barang)
    elif pilihan_menu == "Beli Barang":
        beli_barang(barang)
    elif pilihan_menu == "Modifikasi Barang":
        modifikasi_barang(barang)
    elif pilihan_menu == "Keluar":
        st.warning("Keluar dari aplikasi.")

# Main function untuk menjalankan aplikasi
def main():
    st.title("Aplikasi Management Kasir")
    st.subheader('_dibuat Oleh_ :red[Muh]:blue[fhri] is')
    st.write("Selamat datang di aplikasi manajemen barang. Pilih menu di samping untuk memulai.")
    
    barang = baca_dari_excel("data_barang.xlsx")  # Baca data barang dari file Excel
    
    # Simpan data barang dalam session state
    if 'barang' not in st.session_state:
        st.session_state.barang = barang
    
    tampilkan_menu(st.session_state.barang)  # Tampilkan menu utama

if __name__ == "__main__":
    main()
